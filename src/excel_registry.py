"""
Registro persistente de actas E-14 enviadas a Azure Blob Storage.

Evita re-descargar y re-subir PDFs ya procesados en ejecuciones anteriores.
Usa un .xlsx para que sea inspeccionable sin herramientas especiales.
"""
import os

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

COLUMNS = [
    "clave",
    "departamento_codigo",
    "departamento_nombre",
    "municipio_codigo",
    "municipio_nombre",
    "zona_codigo",
    "puesto_codigo",
    "puesto_nombre",
    "mesa_numero",
    "corporacion_acronimo",
    "corporacion_nombre",
    "url_descarga",
    "sha256",
    "tamano_bytes",
    "numero_paginas",
    "hash_archivo_original",
    "container",
    "blob_nombre",
    "blob_url",
    "blob_metadata_url",
    "fecha_descarga_utc",
    "fecha_subida_utc",
]


def make_key(depto, municipio, zona, puesto, mesa):
    return f"{depto}|{municipio}|{zona}|{puesto}|{mesa}"


def load_registry(path: str):
    """Devuelve (set_de_claves, workbook). Crea el archivo si no existe."""
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        keys = set()
        try:
            key_col = header.index("clave")
        except ValueError:
            return set(), wb
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[key_col]:
                keys.add(row[key_col])
        return keys, wb

    wb = Workbook()
    ws = wb.active
    ws.title = "actas_enviadas"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    return set(), wb


def is_uploaded(registry_keys: set, depto, municipio, zona, puesto, mesa) -> bool:
    return make_key(depto, municipio, zona, puesto, mesa) in registry_keys


def append_record(path: str, registry_keys: set, wb, record: dict):
    """Agrega una fila al registro y guarda a disco. No-op si la clave ya existe."""
    key = make_key(
        record.get("departamento_codigo", ""),
        record.get("municipio_codigo", ""),
        record.get("zona_codigo", ""),
        record.get("puesto_codigo", ""),
        record.get("mesa_numero", ""),
    )
    if key in registry_keys:
        return
    ws = wb.active
    row = [record.get(col) for col in COLUMNS]
    row[COLUMNS.index("clave")] = key
    ws.append(row)
    registry_keys.add(key)
    parent = os.path.dirname(os.path.abspath(path))
    os.makedirs(parent, exist_ok=True)
    wb.save(path)


def merge_registries(source_paths: list, dest_path: str):
    """Fusiona varios registros de workers en el registro maestro. Ignora duplicados.

    Usado por scraper_parallel.py al final de cada run para consolidar los
    archivos por-worker en un unico Excel persistente entre iteraciones.
    """
    dest_keys, dest_wb = load_registry(dest_path)
    dest_ws = dest_wb.active
    added = 0
    for path in source_paths:
        if not os.path.exists(path):
            continue
        _, src_wb = load_registry(path)
        src_ws = src_wb.active
        header = [cell.value for cell in src_ws[1]]
        for row in src_ws.iter_rows(min_row=2, values_only=True):
            record = dict(zip(header, row))
            key = record.get("clave")
            if key and key not in dest_keys:
                dest_ws.append([record.get(col) for col in COLUMNS])
                dest_keys.add(key)
                added += 1
    if added > 0:
        parent = os.path.dirname(os.path.abspath(dest_path))
        os.makedirs(parent, exist_ok=True)
        dest_wb.save(dest_path)
    return added
